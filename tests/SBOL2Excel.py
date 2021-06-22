#wasderivedfrom: source
#remove identity, persistenID, displayID, version
#remove attachment (if empty)
#add library sheets
#add postprocessing function to remove unecessaries

import sbol2
import pandas as pd 
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side
from requests_html import HTMLSession

class seqFile:
    
#     def __init__(self, homeSpace, document, sheet):
    def __init__(self):
        #global varibales for homespace and document 
        self.homeSpace = 'http://sys-bio.org'
        self.document = 'pichia_toolkit_KWK_v002.xml'
        self.sheet = 'ontologies.xlsx'
        
    def roleVars(self):
        df = pd.read_excel('../resources/templates/' + self.sheet, index_col = 0, sheet_name = 1,usecols = [1,2])
        role_convert_dict = df.to_dict()
        role_name = role_convert_dict['URI']
        final = {uri:role for role, uri in role_name.items()}
        return final
    
    def orgVars(self):
        df = pd.read_excel('../resources/templates/' + self.sheet, index_col = 0, sheet_name = 2,usecols = [0,1])
        organism_convert_dict = df.to_dict()
        organism_name = organism_convert_dict['txid']
        final = {str(txid):organism for organism, txid in organism_name.items()}
        return final
        
    def inspectDocInfo(self):
        #declare homespace 
        sbol2.setHomespace(self.homeSpace)
        doc = sbol2.Document()
        doc.read('../resources/templates/' + self.document)
#         doc.read(self.document)
        #print document information 
        print(doc) 
    
    def printDocContents(self):
        #declare homespace 
        sbol2.setHomespace(self.homeSpace)
        doc = sbol2.Document()
        doc.read('../resources/templates/' + self.document)
#         doc.read(self.document)
        #print document contents
        for obj in doc:
            print(obj)
    
    def readDocChart(self):
        #declare homespace 
        sbol2.setHomespace(self.homeSpace)
        doc = sbol2.Document()
        doc.read('../resources/templates/' + self.document)
#         doc.read(self.document)
        #create a dictionary to hold all the component defintions' information 
        componentDefinitions = {}
        #iterate through the component definitions 
        roleDict = self.roleVars()
        orgDict = self.orgVars()
        for cd in doc.componentDefinitions:
            cdType = cd.type
            #create a dictionary that has a key for the component definition's identity, 
            #and a value for all of its features
            componentFeatures = {}
            persistentIdentity = cd.properties['http://sbols.org/v2#persistentIdentity'][0]
            #iterate through the properties of the component defintions and set them equal to propValue variable 
            for prop in cd.properties:
                try:
                    propValue = cd.properties[prop][0]
                except:
                    propValue = cd.properties[prop]
                    #extract attribute property type
                if propValue == []:
                    propValue = ''
                idx = prop.find('#')
                #if parsing conditions meet, append them into the componentFeatures dictionary as necessary
                if idx >= 1:
                    prop = prop[idx+1:]
                if prop == 'type':
                    prop = 'types'
                if prop == 'http://purl.org/dc/terms/title':
                    prop = 'title'
                if prop == 'http://purl.org/dc/terms/description':
                    prop = 'description'
                if prop == 'http://purl.obolibrary.org/obo/OBI_0001617':
                    prop = 'OBI_0001617'
                propValue = columnMethods(prop, propValue, doc, cdType, roleDict, orgDict).colV
                componentFeatures[prop]= str(propValue)
            #append each componentFeatures dictionary as a value into the componentDefinitions 
            #dictionary with the 'persistentIdentity' serving as the key
            componentDefinitions[persistentIdentity] = componentFeatures
        #return the dictionary of information (temporary, maybe return true if read in correctly)
        return componentDefinitions

    def TEMP_readDocChart1(self):
        arr2 = ['Part Name', 
                'Role', 
                'Design Notes',
               'Altered Sequence',
               'Part Description',
               'Data Source Prefix',
               'Data Source',
               'Source Organism',
               'Target Organism',
               'Circular',
               'length (bp)',
               'Sequence',
               'Data Source',
               'Composite']
        exc = self.readDocChart()
        df = pd.DataFrame.from_dict(exc, orient = "index")
        inputList = set(df)
        print(inputList)
        columnOrder = set(arr2)
        print(columnOrder)
        differ = inputList.difference(columnOrder)
        intersection = inputList.intersection(columnOrder)
        final = list(intersection) + list(differ)
        print(final)
        x = df[final].to_dict()
        dframe = pd.DataFrame.from_dict(x, orient = "index")
        return dframe.T
    
    def displayDocChart(self):
        disp = self.readDocChart()
        return pd.DataFrame.from_dict(disp, orient = "index")
    
    def column_string(self, n):
        string = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string
    
    def returnExcelChart(self):
        wb = load_workbook('../resources/templates/Book4.xlsx')
        ws = wb.active
        exc = self.readDocChart()
        df = pd.DataFrame.from_dict(exc, orient = "index")
        ft1 = Font(name = 'Arial', size = 12, color = '548235')
        ft2 = Font(name = 'Calibri', size = 11, bold = True)
        hold = dataframe_to_rows(df, index=False, header=True)
        counter = 0
        ws['A18'].value = ''
        for r in hold:
            if r == [None]:
                continue
            ws.append(r)
            counter += 1
        tab = Table(displayName="Parts_Lib", ref=f"A19:{self.column_string(len(df.columns))}{(len(df) * 2) - 2}")
        style = TableStyleInfo(name="TableStyleLight7", showFirstColumn=False, 
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        cellColor = PatternFill(patternType ='solid', 
                                fgColor = 'DDEBF7')
        cellBorder = Side(border_style='medium', color="000000")
        
        x = df.columns
        cellIndex = len(x)
        for n in range (65, 65 + len(x)):
            alpha = chr(n)
            ws[f'{alpha}19'].fill = cellColor  
            ws[f'{alpha}19'].border = Border(top= cellBorder)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        counter = 0
        for n in range(len(df) - 1, (len(df) * 2 - 1)):
            counter = counter + 1
            for i in ws[n]:
                i.font = ft1
        titleFormat = len(df)
        if titleFormat % 2 > 0:
            titleFormat = titleFormat - 1
        for j in range(19, titleFormat):
            for x in ws[j]:
                x.font = ft2
        wb.save('../outputs/sequenceparts.xlsx')
        wb.close()
    
class columnMethods:
    
    def __init__(self, colN, colV, doc, cdType, roleDict, orgDict):
        self.colN = colN
        self.colV = colV
        self.doc = doc
        self.cdType = cdType
        self.roleDict = roleDict
        self.orgDict = orgDict
        try:
            return getattr(self, self.colN)()
        except AttributeError:
            return getattr(self, 'no_change')()
        
    def no_change(self):
        pass
    
    def role(self):   
        roleVal = str(self.colV)
        if roleVal in self.roleDict:
            self.colV = self.roleDict[roleVal]
            
    def types(self):
         self.colV = self.colV.split('#')[-1]
            
    def sequence(self):
        self.colV = self.doc.getSequence(self.colV).elements
        
    def sourceOrganism(self):
        orgVal = str(self.colV)
        orgVal = orgVal.split('=')[-1]
        txid = self.colV.split('=')[-1]
        if orgVal in self.orgDict:
            self.colV = self.orgDict[orgVal] 
        else:
            session = HTMLSession()
            r = session.get(self.colV)
            v = r.html.find('strong', first = True)
            self.colV = v.text
            self.orgDict[txid] = self.colV

    def targetOrganism(self):
        orgVal = str(self.colV)
        orgVal = orgVal.split('=')[-1]
        txid = self.colV.split('=')[-1]
        if orgVal in self.orgDict:
            self.colV = self.orgDict[orgVal]
        else:
            session = HTMLSession()
            r = session.get(self.colV)
            v = r.html.find('strong', first = True)
            self.colV = v.text
            self.orgDict[txid] = self.colV