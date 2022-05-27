import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from rdflib import Graph
import logging
import sbol2excel.helper_functions as hf
import sbol2excel.column_methods as cm


def process_col_val(ws, start_row, len_col, role_dict, org_dict):
    """Create Excel Hyperlink."""
    for column in ws.iter_cols(min_col=0, max_col=len_col):
        col_name = column[start_row].value
        for cell in column:
            if cell.value is col_name or cell.value is None:
                continue
            else:
                cm.col_methods(ws, cell.coordinate, col_name, str(cell.value), role_dict,
                                org_dict).prop_val


def sbol_to_df(sbol_doc_path, role_dict, org_dict):
    """Utilize RDFLib to collect document contents."""
    g = Graph()
    sbol_doc_path = str(sbol_doc_path)
    print(sbol_doc_path)
    g.parse(sbol_doc_path)
    data_dict = {}
    for index, (subject, predicate, _object) in enumerate(g):
        # collect subject, predicate, and object triples
        subject = str(subject)
        predicate = str(predicate)
        _object = str(_object)
        # print(subject)
        # create dataframe the prepares the triples to be output to excel
        if subject in data_dict:
            if predicate in data_dict[subject]:
                data_dict[subject][predicate] = data_dict[subject][predicate] + ", " + _object
            else:
                data_dict[subject][predicate] = _object
        else:
            data_dict[subject] = {predicate: _object}
    return data_dict


def df_to_excel(df, output_path, output_template, role_dict, org_dict):
    """Output a df into an excel template.

    Args:
        df (pandas dataframe): The dataframe to put into the excel
                spreadsheet
        output_path (file path): Path to where the new excel file should
                be output
        output_template (string): the name of the output template to use

    Raises:
        TypeError: if df is not a pandas dataframe
        ValueError: if the output_template is not a valid name
    """
    # input type checking
    if isinstance(df, pd.core.frame.DataFrame) is False:
        # if type(df) is not pd.core.frame.DataFrame:
        raise TypeError

    # first row to put data into (one below last filled row)
    start_row = 1
    # first cell to put data into, by default the column data is put into is A
    up_l_tbl_cell = f"A{start_row}"

    # df_num_rows = len(df)
    # df_num_cols = len(df.columns)
    # bt_r_tbl_cell = f"{hf.col_to_num(df_num_cols)}{df_num_rows+start_row}"

    # load workbook

    file_dir = os.path.dirname(__file__)
    SBOL2Excel_path = os.path.split(file_dir)[0]
    output_template_path = os.path.join(SBOL2Excel_path, 'sbol2excel',
                                        'Output_Templates', output_template)
    if not os.path.isfile(output_template_path):
        raise ValueError
    wb = load_workbook(output_template_path)
    ws = wb.active
    df_row_obj = dataframe_to_rows(df, index=False, header=True)

    # empty row added to maintain space between previous full row and
    # the table of parts inserted
    ws.append([])
    # add df information

    rdf_types = df['rdf:type'].unique().tolist()
    rdf_df = {}
    # print(df.columns)
    # for i in df.columns:
    #     print(i)
    #     if isinstance(i, str) is True:
    #         print(i)
    #     print('********')

    for type in rdf_types:
        temp = df.loc[df['rdf:type'] == type].replace(r'^s*$', float('NaN'), regex = True)
        temp = temp.dropna(axis=1, how='all').reset_index(drop=True)
        temp = temp.drop(columns=['rdf:type'])
        rdf_df[type] = temp

    for sheet_num in range(0, len(rdf_types)):
        if rdf_types[sheet_num] and rdf_types[sheet_num] in rdf_df:
            ws = wb.create_sheet(str(rdf_types[sheet_num].split('#')[-1]), sheet_num)
            out_frame_df = pd.DataFrame(rdf_df[rdf_types[sheet_num]])
            out_frame_obj = dataframe_to_rows(out_frame_df, index=False, header=True)
            for row in out_frame_obj:
                ws.append(row)
    # ! hol up
    # for row in df_row_obj:
    #     ws.append(row)
    # ! hol up

    # process_col_val(ws, start_row-1, len(df.columns), role_dict, org_dict)

        df_num_rows = len(out_frame_df)
        df_num_cols = len(out_frame_df.columns)
        bt_r_tbl_cell = f"{hf.col_to_num(df_num_cols)}{df_num_rows+start_row}"

        tab = Table(displayName=f"Table{sheet_num}", ref=f"{up_l_tbl_cell}:{bt_r_tbl_cell}")

        # style the table
        style = TableStyleInfo(name="TableStyleLight4",
                            showFirstColumn=False,
                            showLastColumn=False,
                            showRowStripes=True,
                            showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        # freeze header row panes
        freeze = ws[f"A{start_row+1}"]
        ws.freeze_panes = freeze

    # Save
    wb.save(output_path)
    logging.warning(f'Your converted file has been output at {output_path}')
