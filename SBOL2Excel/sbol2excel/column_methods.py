"""Class utilized to process column values."""
from multiprocessing.sharedctypes import Value
from tkinter.tix import Tree
from requests_html import HTMLSession
# import sbol2
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
from openpyxl.styles import NamedStyle
import validators


class col_methods:
    """Class used to carry out a switch case statement.

    Done for different properties in SBOL files
    """

    # def __init__(self, ws, prop_nm, prop_val, role_dict, org_dict):
    def __init__(self, ws, addy, prop_nm, prop_val, role_dict, org_dict):
        """Switch statement to call different methods based on prop_nm.

        Args:
            prop_nm (str): the name of the property
            prop_val (str): the value of the property
            sbol_doc (sbol document): sbol document containing the properties
                                being passed in
            role_dict (dictionary): maps sequence ontology terms to human
                                readable names
            org_dict (dictionary): maps ncbi txids to human readable names

        """
        # global varibales for dataframe switch statements

        self.ws = ws
        self.addy = addy
        self.prop_nm = prop_nm
        self.prop_val = prop_val
        self.role_dict = role_dict
        self.org_dict = org_dict

        function_call_dict = {
            # 'sbol:persistentIdentity': 'hyperlink_end_idx',
            # 'prov:wasGeneratedBy': 'hyperlink_end_idx',
            # 'sbol:role': 'role',
            # 'sbh:ownedBy': 'hyperlink_end_idx',
            # 'sbol:type': 'hyperlink_end_idx',
            'rdf:type': 'hyperlink_end_idx',
            # 'sbol:orientation': 'hyperlink_end_idx',
            # 'sbol:definition': 'hyperlink_end_idx',
            # 'sbol:functionalComponent': 'hyperlink_end_idx',
            # 'sbol:access': 'hyperlink_end_idx',
            # 'sbh:topLevel': 'hyperlink_end_idx',
            # 'sbol:sequence': 'hyperlink_end_idx',
            # 'sbol:interaction': 'hyperlink_end_idx',
            # 'sbol:direction': 'hyperlink_end_idx',
            # 'sbol:participant': 'hyperlink_end_idx',
            # 'sbol:source': 'hyperlink_end_idx',
            # 'prov:qualifiedUsage': 'hyperlink_end_idx',
            # 'prov:qualifiedAssociation': 'hyperlink_end_idx',
            # 'sbol:participation': 'hyperlink_end_idx',
            'dcterms:created': 'date_created',
            'sbol:format': 'form'
            # ! {WARNING: Crucial key:value pairs
            # TODO: Create the right prop_val processors for columns
            #   'http://sbols.org/v2#encoding': 'sequence',
            #   'rdf:type': 'types',
            #   'sbol:type': 'types',
            #   'https://identifiers.org/taxonomy': 'organism'
            }
        # ! WARNING: Crucial key:value pairs}
        if self.prop_nm in function_call_dict:
            self.prop_nm = function_call_dict[self.prop_nm]
        else:
            self.prop_nm = 'hyperlink_end_idx'
        # if the column name matches the function name, call the function
        try:
            getattr(self, self.prop_nm)()
        # if the column name does not match the function name, call 'no_change'
        except AttributeError:
            getattr(self, 'no_change')()

    def no_change(self):
        """Else case for the switch statement."""
        pass

    def link_to_value(self, link, value):
        """Asssign hyperlink to processed value."""
        self.ws[self.addy].hyperlink = link
        self.ws[self.addy].value = value
        self.ws[self.addy].style = "Hyperlink"

    def check_end_idx(self, value):
        """Check the process prop name for correct value."""
        if len(value) > 1:
            try:
                int(value[-1])
                return value[-2]
                # return value[0]
            except ValueError:
                return value[-1]
        else:
            pass

    def multi_processor(self, value):
        """Process multi-value cells."""
        if isinstance(value, int) is True:
            _list = value.split(',')
            value_holder = []
            for uri in _list:
                check_list = uri.split('/')
                value = self.check_end_idx(check_list)
                value_holder.append(value)
            self.ws[self.addy].value = ', '.join(value_holder)
        else:
            pass

    def num_val_check(self, value):
        """Process multi-value cells."""
        _list = value.split(',')
        self.multi_processor(_list)

    def hyperlink_end_idx(self):
        """Create a hyperlink to be output to Excel."""
        # valid = validators.url(self.prop_val)
        # TODO: try a check with isinstance(self.prop_val, nan)
        if isinstance(self.prop_val, str):
            if '#' in self.prop_val:
                try:
                    prop_val_split = self.prop_val.split('#')
                    value = self.check_end_idx(prop_val_split)
                    self.link_to_value(self.prop_val, value)
                except ValueError:
                    prop_val_split = self.prop_val.split('/')
                    print(prop_val_split)
                    # value = self.check_end_idx(prop_val_split)
                    # self.link_to_value(self.prop_val, value)
            elif ',' in self.prop_val:
                self.multi_processor(self.prop_val)
            # prop_val_split = self.prop_val.split('/')
            # value = self.check_end_idx(prop_val_split)
            # self.link_to_value(self.prop_val, value)
        else:
            self.no_change()
            #     else:
        #         prop_val_split = self.prop_val.split('/')
        #         value = self.check_end_idx(prop_val_split)
        #         self.link_to_value(self.prop_val, value)
        # ! IMPORTANT
        #     else:
        #         prop_val_split = self.prop_val.split('/')
        #         value = self.check_end_idx(prop_val_split)
        #         self.link_to_value(self.prop_val, value)
        # elif ',' in self.prop_val:
        #     self.multi_processor(self.prop_val)
        # else:
        #     self.no_change()
        # ! IMPORTANT

    # TODO: is there a smarter way to do this check?
    def date_created(self):
        """Process the date created terms."""
        if self.prop_val != 'nan':
            value = self.prop_val
            value = self.prop_val.replace('T', ' ')
            value = value.replace('Z', '')
            self.ws[self.addy] = value
        else:
            self.ws[self.addy] = None

    def form(self):
        """Process sbol:format."""
        valid = validators.url(self.prop_val)
        if valid:
            value = self.prop_val.split('/')[-1] + ", " + self.prop_val.split('/')[-2]
            self.link_to_value(self.prop_val, value)
        else:
            pass
            # value = self.prop_val.split('/')[-1] + ", " + self.prop_val.split('/')[-2]

    def role(self):
        """Utilize prop_val as the key in a dictionary to get the new value.

        It is a way of converting an ontology term to a human readable one
        """
        # role_val = str(self.prop_val)
        role_val = self.prop_val
        if role_val is None:
            pass
        elif role_val in self.role_dict:
            self.prop_val = self.role_dict[role_val]
            self.link_to_value(role_val, self.prop_val)

    # ! {WARNING: Crucial Functions; Make sure to uncomment
    # def types(self):
    #     """Split types uri to only be the last bit after the final hash.

    #     Raises:
    #         ValueError: If self.prop_val does not contain a #
    #     """
    #     self.prop_val = self.prop_val
    #     if '#' not in self.prop_val:
    #         pass
    #     else:
    #         self.prop_val = self.prop_val.split('#')[-1]

    # def sequence(self):
    #     """Get the sequence from the document based on the sequence uri.

    #     Raises:
    #         TypeError: If the prop_val from initialisation is not a uri
    #                     or string
    #         ValueError: If the prop_val from initialisation is not a uri
    #                 in the sbol document provided at initialisation
    #     """
    #     try:
    #         temp = self.sbol_doc.getSequence(self.prop_val)
    #         self.prop_val = temp.elements
    #     except sbol2.sbolerror.SBOLError:
    #         # if uri not found in document
    #         raise ValueError

    # def organism(self):
    #     """Convert a uri containing a txid into a human readable name.

    #     Done either by using the ontology provided or by pulling the name from
    #     the ncbi database. If the name is pulled from the database it is added
    #     to the ontology for the rest of the program run (the assumption is
    #     that a rare organism may be used multiple times)

    #     Raises:
    #         TypeError: if self.prop_val is not a string or uri
    #         ValueError: if self.prop_val doesn't contain
    #                     'https://identifiers.org/taxonomy:'
    #         TypeError: if self.org_dict is not a dictionary
    #     """
    #     if 'https://identifiers.org/taxonomy:' not in self.prop_val:
    #         raise ValueError
    #     if type(self.org_dict) is not dict:
    #         raise TypeError

    #     # txid = str(self.prop_val).split(':')[-1]
    #     txid = self.prop_val.split(':')[-1]

    #     if txid in self.org_dict:
    #         self.prop_val = self.org_dict[txid]
    #     else:
    #         session = HTMLSession()
    #         r = session.get(f'https://identifiers.org/taxonomy:{txid}')
    #         v = r.html.find('strong', first=True)
    #         self.prop_val = v.text
    #         self.org_dict[txid] = self.prop_val
    # ! WARNING: Crucial Functions; Make sure to uncomment}
