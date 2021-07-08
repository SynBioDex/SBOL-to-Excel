import sbol2
import pandas as pd
import os
import logging
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side
from requests_html import HTMLSession

#wasderivedfrom: source
#remove identity, persistenID, displayID, version
#remove attachment (if empty)
#add library sheets
#add postprocessing function to remove unecessaries


class seqFile:

    def __init__(self, file_path_in, output_path):
        # global varibales for homespace, document, and sheet
        self.homeSpace = 'https://sys-bio.org'
        self.document = file_path_in
        self.file_location_path = os.path.dirname(__file__)
        self.sheet = os.path.join(self.file_location_path, 'ontologies.xlsx')
        self.output_template = os.path.join(self.file_location_path, 'Template_to_Output_Into_v001.xlsx')
        self.output_path = output_path

    def roleVariables(self):
        # set Excel file into a dataframe
        df = pd.read_excel(self.sheet, index_col=0,
                           sheet_name=1, usecols=[1, 2])
        # convert the dataframe into a dictionary
        roleConvertDict = df.to_dict()
        # set dictionary indices and values (use column 'URI' in excel sheet)
        roleName = roleConvertDict['URI']
        # switch indices' and values' postions
        roleDictionary = {uri: role for role, uri in roleName.items()}
        return roleDictionary

    def orgVariables(self):
        # set Excel file into a dataframe
        df = pd.read_excel(self.sheet, index_col=0,
                           sheet_name=2, usecols=[0, 1])
        # convert the dataframe into a dictionary
        organismConvertDict = df.to_dict()
        # set dictionary indices and values (use column 'txid' in excel sheet)
        organismName = organismConvertDict['txid']
        # switch indices' and values' postions
        organismDictionary = {str(txid): organism for organism, txid in organismName.items()}
        return organismDictionary

    def overarching(self):
        df = self.readDocChart()
        df = self.reorder_columns(df)
        self.returnExcelChart(df)
        return

    def readDocChart(self):
        # declare homespace
        sbol2.setHomespace(self.homeSpace)
        doc = sbol2.Document()
        doc.read(self.document)
        # create a dictionary to hold all the component defintions' information
        componentDefinitions = {}
        # iterate through the component definitions
        roleDict = self.roleVariables()
        orgDict = self.orgVariables()
        for cd in doc.componentDefinitions:
            cdType = cd.type
            # create a dictionary that has a key for the
            # component definition's identity,
            # and a value for all of its features
            componentFeatures = {}
            persistentIdentity = cd.properties['http://sbols.org/v2#persistentIdentity'][0]
            # iterate through the properties of the component defintions
            # and set them equal to propValue variable
            for prop in cd.properties:
                try:
                    propValue = cd.properties[prop][0]
                except (IndexError):
                    propValue = cd.properties[prop]
                    # extract attribute property type
                if propValue == []:
                    propValue = ''
                prop = self.prop_convert(prop)
                propValue = columnMethods(prop, propValue, doc, cdType,
                                          roleDict, orgDict).colV
                componentFeatures[prop] = str(propValue)
            # append each componentFeatures dictionary as a
            # value into the componentDefinitions
            # dictionary with the 'persistentIdentity' serving as the key
            componentDefinitions[persistentIdentity] = componentFeatures
        # return the dictionary of information (temporary, maybe
        # return true if read in correctly)
        
        doc_chart = pd.DataFrame.from_dict(componentDefinitions, orient="index")
        return doc_chart

    def prop_convert(self, prop):
        if type(prop) is str:
            idx = prop.find('#')
            # if parsing conditions meet, append them into the
            # componentFeatures dictionary as necessary
            if idx >= 1:
                prop = prop[idx + 1:]
            if prop == 'type':
                prop = 'types'
            if prop == 'http://purl.org/dc/terms/title':
                prop = 'title'
            if prop == 'http://purl.org/dc/terms/description':
                prop = 'description'
            if prop == 'http://purl.obolibrary.org/obo/OBI_0001617':
                prop = 'OBI_0001617'
            return (prop)
        else:
            raise ValueError()
    
    def displayDocChart(self):
        #display the dataframe
        return pd.DataFrame.from_dict(self.readDocChart(), orient = "index")

    # def TEMP_readDocChart1(self):
    #     # demo of table column names
    #     columnNames = ['Part Name',
    #                    'Role',
    #                    'Design Notes',
    #                    'Altered Sequence',
    #                    'Part Description',
    #                    'Data Source Prefix',
    #                    'Data Source',
    #                    'Source Organism',
    #                    'Target Organism',
    #                    'Circular',
    #                    'length (bp)',
    #                    'Sequence',
    #                    'Data Source',
    #                    'Composite']
    #     # import dataframe dictionary
    #     # convert dictionary to dataframe
    #     df = self.displayDocChart()
    #     # type caste dataframe to a set
    #     dfSet = set(df)
    #     # type caste column names to a set
    #     columnNameOrder = set(columnNames)
    #     # check difference between the dataframe set and the column name order
    #     dfSetDifference = dfSet.difference(columnNameOrder)
    #     # check intersection between the datframe set and the column name order
    #     dfSetIntersection = dfSet.intersection(columnNameOrder)
    #     # combine the type casted difference and intersection
    #     finalSetList = list(dfSetIntersection) + list(dfSetDifference)
    #     # set list to dictionary
    #     return finalSetList

    def reorder_columns(self, df):
        # demo of table column names
        # columnNames = col_list
        columnNames = ['Part Name',
                       'Role',
                       'Design Notes',
                       'Altered Sequence',
                       'Part Description',
                       'Data Source Prefix',
                       'Data Source',
                       'Source Organism',
                       'Target Organism',
                       'Circular',
                       'length (bp)',
                       'Sequence',
                       'Data Source',
                       'Composite']
        # type caste dataframe to a set
        dfSet = set(df)
        # type caste column names to a set
        columnNameOrder = set(columnNames)
        # check difference between the dataframe set and the column name order
        dfSetDifference = dfSet.difference(columnNameOrder)
        # check intersection between the datframe set and the column name order
        dfSetIntersection = dfSet.intersection(columnNameOrder)
        # combine the type casted difference and intersection
        finalSetList = list(dfSetIntersection) + list(dfSetDifference)
        # set list to dictionary
        return finalSetList

    # def displayDocChart(self):
    #     # display the dataframe
    #     return pd.DataFrame.from_dict(self.readDocChart(), orient="index")

    def columnString(self, n):
        # loop through column length in order to get string appropriate
        # values for excel sheet rows and columns
        string = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string

    def returnExcelChart(self, df):
        start_row = 18
        start_cell = f'A{start_row}'
        # load a workbook
        wb = load_workbook(self.output_template)
        ws = wb.active
        # load raw dataframe to df
        # df = self.readDocChart()
        # set font features
        ft1 = Font(name='Arial', size=12, color='548235')
        ft2 = Font(name='Calibri', size=11, bold=True)
        hold = dataframe_to_rows(df, index=False, header=True)
        # counter = 0
        # loop through worksheet
        ws[start_cell].value = ''
        for r in hold:
            # if a specific cell is empty, continue to loop past it
            if r == [None]:
                continue
            ws.append(r)
            # counter += 1
        # set table features
        tab = Table(displayName="Parts_Lib", ref=f"A{start_row +1}:{self.columnString(len(df.columns))}{(len(df) * 2) - 2}")
        style = TableStyleInfo(name="TableStyleLight7", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True,
                               showColumnStripes=False)
        cellColor = PatternFill(patternType='solid',
                                fgColor='DDEBF7')
        cellBorder = Side(border_style='medium', color="000000")
        # cellIndex = len(x)
        # gives cells within specified range their table attributes
        for col in range(1, len(df.columns) + 1):
            alpha = self.columnString(col)
            ws[f'{alpha}{start_row+1}'].fill = cellColor
            ws[f'{alpha}{start_row+1}'].border = Border(top=cellBorder)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        # counter = 0
        # gives cells within specified range their font attributes
        for row in range(len(df) - 1, (len(df) * 2 - 1)):
            # counter = counter + 1
            for cell in ws[row]:
                cell.font = ft1
        # gives cells within specified range their font attributes
        # (these are special features for the title)
        num_rows = len(df)
        if num_rows % 2 > 0:
            num_rows = num_rows - 1
        for j in range(19, num_rows):
            for x in ws[j]:
                x.font = ft2
        # output the file
        wb.save(self.output_path)
        wb.close()
        logging.warning(f'Your converted file has been output at {self.output_path}')


class columnMethods:

    def __init__(self, colN, colV, doc, cdType, roleDict, orgDict):
        # global varibales for dataframe switch statements
        self.colN = colN
        self.colV = colV
        self.doc = doc
        self.cdType = cdType
        self.roleDict = roleDict
        self.orgDict = orgDict
        # if the column name matches the function name, call the function
        try:
            return getattr(self, self.colN)()
        # if the column name does not match the function name, call 'no_change'
        except AttributeError:
            return getattr(self, 'no_change')()

    def no_change(self):
        pass
    # if the specified column role value is within the role column

    def role(self):
        roleVal = str(self.colV)
        if roleVal in self.roleDict:
            self.colV = self.roleDict[roleVal]

    def types(self):
        self.colV = self.colV.split('#')[-1]

    def sequence(self):
        self.colV = self.doc.getSequence(self.colV).elements

    def sourceOrganism(self):
        orgVal = str(self.colV)
        orgVal = orgVal.split('=')[-1]
        txid = self.colV.split('=')[-1]
        if orgVal in self.orgDict:
            self.colV = self.orgDict[orgVal]
        else:
            session = HTMLSession()
            r = session.get(self.colV)
            v = r.html.find('strong', first=True)
            self.colV = v.text
            self.orgDict[txid] = self.colV

    def targetOrganism(self):
        orgVal = str(self.colV)
        orgVal = orgVal.split('=')[-1]
        txid = self.colV.split('=')[-1]
        if orgVal in self.orgDict:
            self.colV = self.orgDict[orgVal]
        else:
            session = HTMLSession()
            r = session.get(self.colV)
            v = r.html.find('strong', first=True)
            self.colV = v.text
            self.orgDict[txid] = self.colV

