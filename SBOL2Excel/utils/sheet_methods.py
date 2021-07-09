import sbol2
import pandas as pd
import openpyxl as oxl
import logging


def readDocChart():
    # declare homespace
    sbol2.setHomespace(self.homeSpace)
    doc = sbol2.Document()
    doc.read(self.document)
    # create a dictionary to hold all the component defintions' information
    componentDefinitions = {}
    # iterate through the component definitions
    roleDict = self.roleVariables()
    orgDict = self.orgVariables()
    for cd in doc.componentDefinitions:
        cdType = cd.type
        # create a dictionary that has a key for the
        # component definition's identity,
        # and a value for all of its features
        componentFeatures = {}
        persistentIdentity = cd.properties['http://sbols.org/v2#persistentIdentity'][0]
        # iterate through the properties of the component defintions
        # and set them equal to propValue variable
        for prop in cd.properties:
            try:
                propValue = cd.properties[prop][0]
            except (IndexError):
                propValue = cd.properties[prop]
                # extract attribute property type
            if propValue == []:
                propValue = ''
            prop = self.prop_convert(prop)
            propValue = columnMethods(prop, propValue, doc, cdType,
                                        roleDict, orgDict).colV
            componentFeatures[prop] = str(propValue)
        # append each componentFeatures dictionary as a
        # value into the componentDefinitions
        # dictionary with the 'persistentIdentity' serving as the key
        componentDefinitions[persistentIdentity] = componentFeatures
    # return the dictionary of information (temporary, maybe
    # return true if read in correctly)

    doc_chart = pd.DataFrame.from_dict(componentDefinitions, orient="index")
    return doc_chart

def returnExcelChart(df):
    start_row = 18
    start_cell = f'A{start_row}'
    # load a workbook
    wb = oxl.load_workbook(self.output_template)
    ws = wb.active
    # load raw dataframe to df
    # df = self.readDocChart()
    # set font features
    ft1 = oxl.styles.Font(name='Arial', size=12, color='548235')
    ft2 = oxl.styles.Font(name='Calibri', size=11, bold=True)
    hold = oxl.utils.dataframe.dataframe_to_rows(df, index=False, header=True)
    # counter = 0
    # loop through worksheet
    ws[start_cell].value = ''
    for r in hold:
        # if a specific cell is empty, continue to loop past it
        if r == [None]:
            continue
        ws.append(r)
        # counter += 1
    # set table features
    tab = oxl.worksheet.table.Table(displayName="Parts_Lib", ref=f"A{start_row +1}:{self.columnString(len(df.columns))}{(len(df) * 2) - 2}")
    style = oxl.worksheet.table.TableStyleInfo(name="TableStyleLight7",
                                               showFirstColumn=False,
                                               showLastColumn=False,
                                               showRowStripes=True,
                                               showColumnStripes=False)
    cellColor = oxl.styles.PatternFill(patternType='solid',
                                       fgColor='DDEBF7')
    cellBorder = oxl.styles.Side(border_style='medium', color="000000")
    # cellIndex = len(x)
    # gives cells within specified range their table attributes
    for col in range(1, len(df.columns) + 1):
        alpha = self.columnString(col)
        ws[f'{alpha}{start_row+1}'].fill = cellColor
        ws[f'{alpha}{start_row+1}'].border = oxl.styles.Border(top=cellBorder)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    # counter = 0
    # gives cells within specified range their font attributes
    for row in range(len(df) - 1, (len(df) * 2 - 1)):
        # counter = counter + 1
        for cell in ws[row]:
            cell.font = ft1
    # gives cells within specified range their font attributes
    # (these are special features for the title)
    num_rows = len(df)
    if num_rows % 2 > 0:
        num_rows = num_rows - 1
    for j in range(19, num_rows):
        for x in ws[j]:
            x.font = ft2
    # output the file
    wb.save(self.output_path)
    wb.close()
    logging.warning(f'Your converted file has been output at {self.output_path}')